"""Прайс-лист поставщика (CSV или Excel) → снимок контракта v2.

Зачем это существует. Оптовые цены на продукты нигде не опубликованы — ни в
России, ни в ЕС (проверено 29.08 на восьми источниках: отраслевые порталы
прячут таблицы за регистрацией, B2B-площадки не отдают цену по позициям).
Поставщик присылает прайс файлом в почту. Значит вход инструмента для
закупщика — не сайт, а файл.

Колонки не угадываются жёстко: у каждого поставщика свой заголовок. Ищем по
списку синонимов, регистр и пробелы не важны. Не нашли обязательную колонку —
говорим об этом вслух и не собираем снимок молча из половины данных.

CSV читается стандартной библиотекой. Excel — через openpyxl, если он есть;
нет — честно скажем, а не упадём.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

from parsers.cards import to_number
from parsers.snapshot import build_snapshot

COLUMNS = {
    "sku": ("артикул", "код", "sku", "код товара", "парт-номер", "штрихкод"),
    "title": ("наименование", "название", "товар", "продукт", "name", "позиция"),
    "price": ("цена", "стоимость", "прайс", "price", "цена руб", "цена, руб"),
    "unit": ("единица", "ед", "ед изм", "фасовка", "unit"),
    "stock": ("наличие", "остаток", "склад", "в наличии", "stock"),
    "moq": ("моq", "moq", "минимальная партия", "мин партия", "кратность"),
    "country": ("страна", "происхождение", "страна происхождения", "country"),
}

OUT_OF_STOCK = ("нет", "0", "под заказ", "ожидается", "отсутствует", "-")


def _norm(text: str) -> str:
    return re.sub(r"[^\wа-яё ]", " ", str(text or "").lower()).strip()


def map_columns(header: list[str]) -> dict[str, int]:
    """Заголовок файла → какая колонка что значит."""
    found: dict[str, int] = {}
    cleaned = [_norm(h) for h in header]
    for field, synonyms in COLUMNS.items():
        for index, cell in enumerate(cleaned):
            if cell in synonyms or any(cell.startswith(s) for s in synonyms):
                found.setdefault(field, index)
                break
    return found


def rows_from_csv(path: Path) -> list[list[str]]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    dialect = csv.Sniffer().sniff(text[:2000], delimiters=";,\t") \
        if text.strip() else csv.excel
    return [row for row in csv.reader(text.splitlines(), dialect) if any(row)]


def rows_from_xlsx(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel читается через openpyxl, его нет в окружении. "
            "Либо `pip install openpyxl`, либо сохраните прайс как CSV — "
            "результат тот же."
        ) from exc
    sheet = load_workbook(path, read_only=True, data_only=True).active
    return [[("" if c is None else str(c)) for c in row]
            for row in sheet.iter_rows(values_only=True) if any(row)]


def read_rows(path: str | Path) -> list[list[str]]:
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return rows_from_xlsx(path)
    return rows_from_csv(path)


def parse_pricelist(path: str | Path, supplier: str,
                    currency: str = "RUB") -> tuple[list[dict], list[str]]:
    """Файл → (позиции, замечания). Замечания — то, что человек должен знать."""
    rows = read_rows(path)
    notes: list[str] = []
    if not rows:
        return [], ["файл пустой"]

    header, body = rows[0], rows[1:]
    columns = map_columns(header)

    for required in ("title", "price"):
        if required not in columns:
            notes.append(
                f"не нашёл колонку «{required}» среди заголовков: "
                f"{', '.join(h for h in header if h)[:120]}")
    if "title" not in columns or "price" not in columns:
        return [], notes

    if "sku" not in columns:
        notes.append("артикула нет — позиции опознаются по названию, "
                     "сопоставление с другими поставщиками будет приблизительным")

    items: list[dict] = []
    for row in body:
        def cell(field: str) -> str:
            index = columns.get(field)
            return row[index].strip() if index is not None and index < len(row) else ""

        title = cell("title")
        if not title:
            continue
        price = to_number(cell("price"))
        stock = _norm(cell("stock"))
        items.append({
            "sku": cell("sku") or re.sub(r"\s+", "-", _norm(title))[:60],
            "shop": supplier,
            "title": title,
            "price": price,
            "currency": currency,
            "price_status": "listed" if price is not None else "unknown",
            "item_status": "ok" if price is not None else "not_found",
            "in_stock": bool(stock) and stock not in OUT_OF_STOCK,
            "unit": cell("unit"),
            "moq": cell("moq"),
            "country": cell("country"),
        })

    if not items:
        notes.append("строк с наименованием не нашлось — возможно, "
                     "заголовок не в первой строке файла")
    return items, notes


def snapshot_from_pricelist(path: str | Path, supplier: str,
                            currency: str = "RUB") -> tuple[dict, list[str]]:
    items, notes = parse_pricelist(path, supplier, currency)
    status = "ok" if items else "unreachable"
    return build_snapshot(supplier, items, status=status), notes
